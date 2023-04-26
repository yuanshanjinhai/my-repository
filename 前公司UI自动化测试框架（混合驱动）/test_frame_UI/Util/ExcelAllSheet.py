# coding=utf-8
from openpyxl import Workbook,load_workbook
import string
from Util.FormatTime import date_time

class ExcelAllSheet():
    def __init__(self,excel_path):
        self.excel_path=excel_path
        self.wb=load_workbook(excel_path)
        # 以下为获取多个列的列号组成的列表，不论列是否有数据
        uper_str = string.ascii_uppercase #生成所有大写字母
        self.col_no_list = list(map(lambda x: x, uper_str)) # 生成基础列号列表，即["A","B"..."Z"]
        for iA in "ABCD": # 在基础列号列表后面加上更多的列，即["A"..."Z"；"AA","AB"..."AZ"；"BA","BB"..."BZ"..."DZ"]
            for iu in uper_str:
                self.col_no_list.append(iA + iu)

    # 获取所有sheet，返回所有sheet名称的列表
    def get_SheetList(self):
        sheet_list = self.wb.sheetnames
        return sheet_list

    # 获取sheet中的最大的行数
    def get_max_row(self,sheet_name):
        ws = self.wb[sheet_name]
        return ws.max_row

    # 获取sheet中的最大列数
    def get_max_col(self,sheet_name):
        ws = self.wb[sheet_name]
        sol_no=ws.max_column # 先获得列号
        return list(filter(lambda x:x==self.col_no_list[sol_no-1],self.col_no_list))[0] # 再转成字母

    # 获取sheet的的最小（起始）行号
    def get_min_row(self,sheet_name):
        ws = self.wb[sheet_name]
        return ws.min_row

    # 获取sheet的最小（起始）列号
    def get_min_col(self,sheet_name):
        ws = self.wb[sheet_name]
        sol_no=ws.min_column # 先获得列号
        return list(filter(lambda x:x==self.col_no_list[sol_no-1],self.col_no_list))[0] # 再转成字母

    # 获取sheet的某一个单元格的值,index为坐标，如"A1","C5"
    def get_value(self,sheet_name,index):
        ws = self.wb[sheet_name]
        return str(ws[index].value).strip()

    def get_value_n(self,sheet_name,index):
        ws = self.wb[sheet_name]
        result=str(ws[index].value)
        if result != "None":
            return result.strip()
        else:
            return "n"

    # 通过坐标写入内容，index为坐标，如"A1"，"F7"，content为写入的内容，但不保存
    def write_content(self,sheet_name,index,content):
        ws = self.wb[sheet_name]
        ws[index]=content

    # 通过坐标写入内容，index为坐标，如"A1"，"F7"，content为写入的内容，写入后自动保存
    def write_content_save(self,sheet_name,index,content):
        ws = self.wb[sheet_name]
        ws[index] = content
        self.wb.save(self.excel_path)

    # 通过坐标写入当前日期
    def write_datetime_en(self,sheet_name,index):
        ws = self.wb[sheet_name]
        ws[index]=date_time()

    # 保存，和write_content和write_datetime配合使用
    def save_content(self):
        self.wb.save(self.excel_path)

# if __name__=='__main__':
#     excel_path=r"D:\pycharm\project\test_frame_UI\Kwyword_Data\kd.xlsx"
#     ins=ExcelAllSheet(excel_path)
#     ins.write_content_save("测试用","B10","abc")
#     ins.save_content()